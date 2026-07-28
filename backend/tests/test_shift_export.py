from __future__ import annotations

import io

from openpyxl import load_workbook

from app.services.shift_service import export_shift_xlsx


def _cycle(caex_id: str, carguio_id: str, hora: int, tonelaje: int) -> dict:
    return {
        "id": f"{caex_id}-{carguio_id}-{hora}",
        "datetime": f"2026-07-04T{hora:02d}:00",
        "fecha_dia": "2026-07-04",
        "turno_calc": "DIA",
        "hora": hora,
        "caex_id": caex_id,
        "camion_modelo": "CAT793F",
        "carguio_id": carguio_id,
        "pala_modelo": "PC5500",
        "material": "MINERAL",
        "origen": "F01",
        "destino": "CHANCADO",
        "fase": "F01",
        "tonelaje": tonelaje,
        "viajes": 1,
    }


def _dataset() -> dict:
    return {
        "source": "wenco-sql-live",
        "today": "2026-07-04",
        "plan": [],
        "cycles": [
            _cycle("CA0001", "EX3600", 8, 220),
            _cycle("CA0002", "EX3600", 9, 210),
            _cycle("CA0001", "EX3470", 10, 200),
        ],
    }


def test_export_shift_xlsx_shift_sheet_has_real_headers():
    payload = export_shift_xlsx(_dataset(), export_type="shift")
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active
    assert sheet.title == "Turno"
    header = [cell.value for cell in sheet[1]]
    assert header == ["fecha", "turno", "toneladas", "meta", "cumplimiento_pct", "brecha", "resumen"]
    assert sheet.max_row >= 2


def test_export_shift_xlsx_caex_sheet_lists_real_trucks():
    payload = export_shift_xlsx(_dataset(), export_type="caex")
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active
    assert sheet.title == "Ranking CAEX"
    header = [cell.value for cell in sheet[1]]
    assert header == ["rank", "caex_id", "modelo", "estado", "toneladas", "ciclos", "rendimiento_tph"]
    caex_ids = {row[1].value for row in sheet.iter_rows(min_row=2)}
    assert caex_ids == {"CA0001", "CA0002"}


def test_export_shift_xlsx_loading_sheet_lists_real_loaders():
    payload = export_shift_xlsx(_dataset(), export_type="loading")
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active
    assert sheet.title == "Unidades de carguio"
    header = [cell.value for cell in sheet[1]]
    assert header == ["carguio_id", "modelo", "estado", "toneladas", "ciclos", "rendimiento_tph", "ubicacion"]
    loader_ids = {row[0].value for row in sheet.iter_rows(min_row=2)}
    assert loader_ids == {"EX3600", "EX3470"}
