from __future__ import annotations

import email
import hashlib
import imaplib
import io
import json
import os
import sqlite3
import unicodedata
import zipfile
from datetime import datetime, timedelta
from email.header import decode_header
from pathlib import Path
from typing import Any

from app.core.distributed import sync_lock

# Ingesta de reportes de averias en Excel (XLS/XLSX) hacia NORTHMINE.
# Fuentes soportadas:
#   1. Carga manual desde el modulo Averias (POST /api/averias/import-xls).
#   2. Carpeta vigilada local (NORTHMINE_AVERIAS_WATCH_DIR): cualquier Excel
#      nuevo que aparezca ahi se importa en el proximo sync.
#   3. Casilla de correo via IMAP (Outlook/Exchange u otro): se conecta,
#      busca correos recientes con adjuntos Excel y los importa.
#      Variables de entorno:
#        NORTHMINE_MAIL_HOST      (default outlook.office365.com)
#        NORTHMINE_MAIL_USER      (casilla, ej: reportes@empresa.cl)
#        NORTHMINE_MAIL_PASSWORD  (contrasena o app password)
#        NORTHMINE_MAIL_FOLDER    (default INBOX)
#        NORTHMINE_MAIL_SENDER    (filtro FROM opcional)
#        NORTHMINE_MAIL_SUBJECT   (filtro SUBJECT opcional)
#        NORTHMINE_MAIL_DAYS      (ventana de busqueda, default 14)
# El parser detecta encabezados por alias (equipo, fecha, inicio, fin,
# duracion, sistema, tipo, descripcion, OT, estado) en cualquier hoja y
# posicion, para tolerar variaciones del formato del reporte.

_DB_PATH = Path(os.getenv("NORTHMINE_AVERIAS_DB", str(Path(__file__).resolve().parents[2] / "northmine_averias.db")))
_ALLOWED_WORKBOOK_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
MAX_WORKBOOK_BYTES = max(1, int(float(os.getenv("NORTHMINE_AVERIAS_MAX_MB", "25")) * 1024 * 1024))
MAX_WORKBOOK_SHEETS = 50
MAX_WORKBOOK_ROWS_PER_SHEET = 100_000
MAX_WORKBOOK_COLUMNS = 200
MAX_XLSX_UNCOMPRESSED_BYTES = max(MAX_WORKBOOK_BYTES * 8, 100 * 1024 * 1024)


def validate_workbook_payload(data: bytes, filename: str) -> None:
    """Reject oversized or structurally suspicious workbook uploads early."""
    safe_name = Path(filename or "").name
    if not safe_name.lower().endswith(_ALLOWED_WORKBOOK_EXTENSIONS):
        raise ValueError("Formato no soportado. Se aceptan .xls, .xlsx y .xlsm.")
    if not data:
        raise ValueError("El archivo esta vacio.")
    if len(data) > MAX_WORKBOOK_BYTES:
        raise ValueError(f"El archivo supera el limite de {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB.")
    if safe_name.lower().endswith((".xlsx", ".xlsm")):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                entries = archive.infolist()
                uncompressed = sum(entry.file_size for entry in entries)
                if len(entries) > 2_000 or uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("El libro excede los limites estructurales permitidos.")
        except zipfile.BadZipFile as exc:
            raise ValueError("El archivo XLSX/XLSM no es un libro valido.") from exc

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "equipment_id": (
        "EQUIPO", "EQUIPOS", "ID EQUIPO", "COD EQUIPO", "CODIGO EQUIPO",
        "CAEX", "CAMION", "UNIDAD", "EQUIPO ID", "MAQUINA", "N EQUIPO",
    ),
    "model": ("MODELO", "TIPO EQUIPO", "FLOTA", "TIPO DE EQUIPO"),
    "fecha": ("FECHA", "DIA", "FECHA EVENTO", "FECHA AVERIA", "FECHA FALLA"),
    "inicio": (
        "INICIO", "HORA INICIO", "FECHA INICIO", "FECHA/HORA INICIO",
        "DESDE", "INICIO DETENCION", "INICIO FALLA", "HORA DETENCION",
    ),
    "fin": (
        "FIN", "TERMINO", "HORA FIN", "FECHA FIN", "FECHA/HORA FIN",
        "HASTA", "FIN DETENCION", "HORA TERMINO", "TERMINO FALLA",
    ),
    "duracion": (
        "DURACION", "DURACION HRS", "DURACION (HRS)", "DURACION HORAS",
        "HORAS", "HRS", "TIEMPO", "TIEMPO DETENCION", "MINUTOS",
        "DURACION MIN", "DURACION (MIN)", "TIEMPO FUERA DE SERVICIO", "TFS",
    ),
    "sistema": ("SISTEMA", "COMPONENTE", "SUBSISTEMA", "CONJUNTO", "AREA FALLA"),
    "tipo": (
        "TIPO", "TIPO FALLA", "TIPO DE FALLA", "TIPO AVERIA", "CLASIFICACION",
        "CATEGORIA", "MODO FALLA", "TIPO MANTENCION", "TIPO DETENCION",
    ),
    "descripcion": (
        "DESCRIPCION", "DETALLE", "OBSERVACION", "OBSERVACIONES", "FALLA",
        "MOTIVO", "COMENTARIO", "COMENTARIOS", "DESCRIPCION FALLA",
        "DESCRIPCION AVERIA", "TRABAJO REALIZADO",
    ),
    "ot": ("OT", "N OT", "NRO OT", "NUMERO OT", "ORDEN TRABAJO", "ORDEN DE TRABAJO", "WO"),
    "estado": ("ESTADO", "STATUS", "SITUACION", "CONDICION"),
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace(".", " ").replace(":", " ").replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS averia_events (
            id TEXT PRIMARY KEY,
            equipment_id TEXT NOT NULL,
            model TEXT,
            fecha TEXT,
            inicio TEXT,
            fin TEXT,
            duracion_min REAL,
            sistema TEXT,
            tipo TEXT,
            descripcion TEXT,
            ot TEXT,
            estado TEXT,
            severidad TEXT,
            origen TEXT NOT NULL,
            source_file TEXT,
            imported_at TEXT NOT NULL,
            extra TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS import_log (
            id TEXT PRIMARY KEY,
            kind TEXT NOT NULL,
            source TEXT,
            filename TEXT,
            events INTEGER,
            imported_at TEXT NOT NULL
        )
        """
    )
    return conn


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]] | None:
    best: tuple[int, dict[int, str]] | None = None
    for row_index, row in enumerate(rows[:20]):
        mapping: dict[int, str] = {}
        for col_index, cell in enumerate(row):
            header = _normalize_header(cell)
            if not header:
                continue
            for field, aliases in _HEADER_ALIASES.items():
                if field in mapping.values():
                    continue
                if header in aliases:
                    mapping[col_index] = field
                    break
        has_core = "equipment_id" in mapping.values()
        if has_core and len(mapping) >= 2:
            if best is None or len(mapping) > len(best[1]):
                best = (row_index, mapping)
    return best


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _cell_str(value)
    if not text:
        return None
    for fmt in (
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d",
        "%H:%M:%S", "%H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _cell_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_str(value)
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _duration_minutes(raw: Any, header_norm: str, inicio: datetime | None, fin: datetime | None) -> float | None:
    number = _cell_number(raw)
    if number is not None:
        if isinstance(raw, datetime):
            number = None
        elif "MIN" in header_norm:
            return round(number, 1)
        elif any(token in header_norm for token in ("HR", "HORA", "TIEMPO", "TFS", "DURACION")):
            # Horas decimales (formato tipico de reportes de mantencion).
            return round(number * 60, 1) if number < 100 else round(number, 1)
        else:
            return round(number, 1)
    if inicio and fin and fin > inicio:
        return round((fin - inicio).total_seconds() / 60, 1)
    return None


def _severity(duracion_min: float | None) -> str:
    if duracion_min is None:
        return "SIN DATO"
    if duracion_min >= 480:
        return "CRITICA"
    if duracion_min >= 120:
        return "ALTA"
    return "MEDIA"


def _load_rows(data: bytes, filename: str) -> list[tuple[str, list[tuple[Any, ...]], list[Any]]]:
    """Devuelve [(sheet_name, rows, header_raw_cells)] para xlsx/xlsm/xls."""
    lower = filename.lower()
    sheets: list[tuple[str, list[tuple[Any, ...]], list[Any]]] = []
    if lower.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        if len(workbook.worksheets) > MAX_WORKBOOK_SHEETS:
            workbook.close()
            raise ValueError("El libro tiene demasiadas hojas.")
        for sheet in workbook.worksheets:
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > MAX_WORKBOOK_ROWS_PER_SHEET or len(row) > MAX_WORKBOOK_COLUMNS:
                    workbook.close()
                    raise ValueError("El libro excede los limites de filas o columnas permitidos.")
                rows.append(tuple(row))
            sheets.append((sheet.title, rows, []))
        workbook.close()
    elif lower.endswith(".xls"):
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise ValueError(
                "El archivo es .xls (Excel 97-2003) y falta la libreria xlrd. "
                "Instala 'xlrd' en el backend o guarda el reporte como .xlsx."
            ) from exc
        book = xlrd.open_workbook(file_contents=data)
        if book.nsheets > MAX_WORKBOOK_SHEETS:
            raise ValueError("El libro tiene demasiadas hojas.")
        for sheet in book.sheets():
            if sheet.nrows > MAX_WORKBOOK_ROWS_PER_SHEET or sheet.ncols > MAX_WORKBOOK_COLUMNS:
                raise ValueError("El libro excede los limites de filas o columnas permitidos.")
            rows = []
            for row_index in range(sheet.nrows):
                values = []
                for col_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, col_index)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        values.append(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode)))
                    else:
                        values.append(cell.value)
                rows.append(tuple(values))
            sheets.append((sheet.name, rows, []))
    else:
        raise ValueError(f"Formato no soportado: {filename}. Se aceptan .xls, .xlsx y .xlsm.")
    return sheets


# --- Formato "ESTATUS DIARIO DE EQUIPOS" (reporte real del correo) ---------
# Libro con una hoja por dia (nombre dd-mm-yyyy). Columnas: TIPO MAQUINA
# (con fill-down), MODELO, N PARQUE, ESTADO (OP / F/S), TIPO MANT, CONCEPTO
# (sistema), DESCRIPCION (componente), DETALLE PARADA, FECHA DETENCION,
# PROYECCION DE ENTREGA, HORAS FUERA DE SERVICIO (texto "1,25 HRS DESDE...").
# El N PARQUE se traduce al ID NORTHMINE: DUMPER 20136 -> CA0136,
# 2916 -> CA2916; PALA/EXCAVADORA 3517 -> EX3517; CARGADOR 3449 -> CF3449.

_ESTATUS_PREFIX = (
    ("DUMPER", "CA"),
    ("PALA", "EX"),
    ("EXCAVADORA", "EX"),
    ("CARGADOR", "CF"),
    ("PERFORADORA", "PF"),
    ("TRACTOR", "TC"),
    ("WHEELDOZER", "WD"),
    ("MOTONIVELADORA", "MN"),
    ("CUBA DE AGUA", "CU"),
    ("CUBA GASOLEO", "CG"),
)


def _estatus_equipment_id(tipo: str, parque: Any) -> str:
    raw = str(parque).strip()
    if raw.endswith(".0"):
        raw = raw[:-2]
    tipo_norm = _normalize_header(tipo)
    prefix = "EQ"
    for key, value in _ESTATUS_PREFIX:
        if key in tipo_norm:
            prefix = value
            break
    if prefix == "CA" and len(raw) == 5 and raw.startswith("2"):
        raw = raw[1:]
    return f"{prefix}{raw}"


def _estatus_hours_minutes(text: str | None) -> float | None:
    if not text:
        return None
    import re

    match = re.search(r"(\d+(?:[.,]\d+)?)", str(text))
    if not match:
        return None
    hours = float(match.group(1).replace(",", "."))
    return round(hours * 60, 1)


def _sheet_report_date(sheet_name: str, rows: list[tuple[Any, ...]]) -> datetime | None:
    # Tanto el nombre de la pestana como la celda FECHA llegan a veces con
    # errores de tipeo (se ha visto '17-06-2026' por 17-07 en ambos lados).
    # Se toman los dos candidatos y gana el mas cercano a las fechas de
    # detencion que contiene la propia hoja.
    candidates: list[datetime] = []
    for row in rows[:4]:
        cell_dates = [cell for cell in row if isinstance(cell, datetime)]
        if cell_dates:
            candidates.append(cell_dates[0])
            break
    try:
        candidates.append(datetime.strptime(sheet_name.strip(), "%d-%m-%Y"))
    except ValueError:
        pass
    unique = list({candidate.date(): candidate for candidate in candidates}.values())
    if not unique:
        return None
    if len(unique) == 1:
        return unique[0]
    body_dates = sorted(
        cell for row in rows[4:] for cell in row if isinstance(cell, datetime)
    )
    if not body_dates:
        return unique[0]
    median = body_dates[len(body_dates) // 2]
    return min(unique, key=lambda candidate: abs((candidate - median).total_seconds()))


def _parse_estatus_sheet(sheet_name: str, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]] | None:
    header_index = None
    columns: dict[str, int] = {}
    for row_index, row in enumerate(rows[:10]):
        norms = {col: _normalize_header(cell) for col, cell in enumerate(row) if cell is not None}
        if not any("PARQUE" in norm for norm in norms.values()):
            continue
        def assign(key: str, col: int) -> None:
            # Primera coincidencia gana: las hojas traen una tabla RESUMEN a la
            # derecha con encabezados repetidos (TIPO MAQUINA, MODELO) que no
            # debe pisar las columnas del detalle.
            if key not in columns:
                columns[key] = col

        for col in sorted(norms):
            norm = norms[col]
            if "PARQUE" in norm:
                assign("parque", col)
            elif norm.startswith("TIPO MAQUINA") and "OCULTAR" not in norm:
                assign("tipo_maquina", col)
            elif norm == "MODELO":
                assign("model", col)
            elif norm == "ESTADO":
                assign("estado", col)
            elif norm.startswith("TIPO MANT"):
                assign("tipo_mant", col)
            elif norm == "CONCEPTO":
                assign("sistema", col)
            elif norm == "DESCRIPCION":
                assign("componente", col)
            elif norm.startswith("DETALLE"):
                assign("detalle", col)
            elif "FECHA DETENCI" in norm:
                assign("inicio", col)
            elif "PROYECCION" in norm:
                assign("proyeccion", col)
            elif "HORAS FUERA" in norm:
                assign("horas", col)
        if "parque" in columns and "estado" in columns:
            header_index = row_index
            break
        columns = {}
    if header_index is None:
        return None

    report_date = _sheet_report_date(sheet_name, rows)
    fecha_reporte = report_date.date() if report_date else None
    events: list[dict[str, Any]] = []
    tipo_maquina = ""

    def col_value(row: tuple[Any, ...], key: str) -> Any:
        col = columns.get(key)
        return row[col] if col is not None and col < len(row) else None

    for row in rows[header_index + 1:]:
        tipo_cell = _cell_str(col_value(row, "tipo_maquina"))
        if tipo_cell:
            tipo_maquina = tipo_cell
        parque = col_value(row, "parque")
        if parque is None or not str(parque).strip():
            continue
        tipo_mant = _cell_str(col_value(row, "tipo_mant"))
        sistema = _cell_str(col_value(row, "sistema"))
        detalle = _cell_str(col_value(row, "detalle"))
        if not tipo_mant and not sistema and not detalle:
            continue  # equipo operativo sin evento
        estado = _cell_str(col_value(row, "estado")) or "OP"
        componente = _cell_str(col_value(row, "componente"))
        inicio = _cell_datetime(col_value(row, "inicio"))
        proyeccion = _cell_datetime(col_value(row, "proyeccion"))
        horas_texto = _cell_str(col_value(row, "horas"))
        duracion = _estatus_hours_minutes(horas_texto)
        dias_detenido = 0
        if estado.upper().replace("/", "") == "FS" and inicio and fecha_reporte:
            dias_detenido = max((fecha_reporte - inicio.date()).days, 0)
            if dias_detenido >= 1:
                duracion = float(dias_detenido * 1440)
        if estado.upper().replace("/", "") == "FS":
            severidad = "CRITICA" if dias_detenido >= 3 else "ALTA"
        else:
            severidad = _severity(duracion)
        tipo_limpio = (tipo_mant or "").split("/")[0].strip() or None
        sistema_limpio = (sistema or "").replace("_", " ").strip() or None
        equipment_id = _estatus_equipment_id(tipo_maquina, parque)
        extra = {
            "tipo_maquina": tipo_maquina or None,
            "componente": componente,
            "estado_equipo": estado,
            "dias_detenido": dias_detenido,
            "proyeccion_entrega": proyeccion.date().isoformat() if proyeccion else None,
            "horas_texto": horas_texto,
        }
        key = f"{equipment_id}|{inicio.isoformat() if inicio else ''}|{detalle or componente or ''}"
        events.append(
            {
                "id": hashlib.sha1(key.encode("utf-8")).hexdigest(),
                "equipment_id": equipment_id,
                "model": _cell_str(col_value(row, "model")),
                "fecha": fecha_reporte.isoformat() if fecha_reporte else None,
                "inicio": inicio.isoformat(timespec="minutes") if inicio else None,
                "fin": None,
                "duracion_min": duracion,
                "sistema": sistema_limpio,
                "tipo": tipo_limpio,
                "descripcion": detalle or componente,
                "ot": None,
                "estado": estado,
                "severidad": severidad,
                "extra": json.dumps(extra, ensure_ascii=False),
                "_report_date": fecha_reporte,
            }
        )
    return events


def parse_workbook(data: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    events: list[dict[str, Any]] = []
    notes: list[str] = []
    estatus_batches: list[tuple[Any, list[dict[str, Any]], str]] = []
    generic_sheets: list[tuple[str, list[tuple[Any, ...]]]] = []
    for sheet_name, rows, _ in _load_rows(data, filename):
        estatus_events = _parse_estatus_sheet(sheet_name, rows)
        if estatus_events is not None:
            report_date = estatus_events[0]["_report_date"] if estatus_events else None
            estatus_batches.append((report_date, estatus_events, sheet_name))
            continue
        generic_sheets.append((sheet_name, rows))

    if estatus_batches:
        # Orden cronologico: la hoja mas reciente actualiza el estado del evento.
        estatus_batches.sort(key=lambda batch: (batch[0] is None, batch[0]))
        for _, batch, sheet_name in estatus_batches:
            for event in batch:
                event.pop("_report_date", None)
            events.extend(batch)
            notes.append(f"Hoja '{sheet_name}': {len(batch)} eventos (formato estatus diario).")

    for sheet_name, rows in generic_sheets:
        found = _find_header_row(rows)
        if not found:
            notes.append(f"Hoja '{sheet_name}': sin encabezados reconocibles, omitida.")
            continue
        header_index, mapping = found
        header_row = rows[header_index]
        header_norms = {col: _normalize_header(header_row[col]) for col in mapping}
        empty_streak = 0
        parsed = 0
        for row in rows[header_index + 1:]:
            values = {field: row[col] if col < len(row) else None for col, field in mapping.items()}
            equipment = _cell_str(values.get("equipment_id"))
            if not equipment:
                empty_streak += 1
                if empty_streak >= 15:
                    break
                continue
            empty_streak = 0
            if _normalize_header(equipment) in _HEADER_ALIASES["equipment_id"]:
                continue  # encabezado repetido
            inicio = _cell_datetime(values.get("inicio"))
            fin = _cell_datetime(values.get("fin"))
            fecha = _cell_datetime(values.get("fecha")) or inicio
            duracion_col = next((col for col, field in mapping.items() if field == "duracion"), None)
            duracion = _duration_minutes(
                values.get("duracion"),
                header_norms.get(duracion_col, "") if duracion_col is not None else "",
                inicio,
                fin,
            )
            extra = {}
            event = {
                "equipment_id": equipment.upper(),
                "model": _cell_str(values.get("model")),
                "fecha": fecha.date().isoformat() if fecha else None,
                "inicio": inicio.isoformat(timespec="minutes") if inicio else None,
                "fin": fin.isoformat(timespec="minutes") if fin else None,
                "duracion_min": duracion,
                "sistema": _cell_str(values.get("sistema")),
                "tipo": _cell_str(values.get("tipo")),
                "descripcion": _cell_str(values.get("descripcion")),
                "ot": _cell_str(values.get("ot")),
                "estado": _cell_str(values.get("estado")),
                "severidad": _severity(duracion),
                "extra": json.dumps(extra, ensure_ascii=False) if extra else None,
            }
            key = "|".join(
                str(event.get(field) or "") for field in ("equipment_id", "fecha", "inicio", "descripcion", "ot")
            )
            event["id"] = hashlib.sha1(key.encode("utf-8")).hexdigest()
            events.append(event)
            parsed += 1
        notes.append(f"Hoja '{sheet_name}': {parsed} eventos (encabezado en fila {header_index + 1}).")
    return events, notes


def import_workbook_bytes(data: bytes, filename: str, origen: str) -> dict[str, Any]:
    validate_workbook_payload(data, filename)
    events, notes = parse_workbook(data, filename)
    imported = 0
    now = datetime.now().isoformat(timespec="seconds")
    updated = 0
    with _connect() as conn:
        for event in events:
            exists = conn.execute("SELECT 1 FROM averia_events WHERE id = ?", (event["id"],)).fetchone()
            if exists:
                conn.execute(
                    """
                    UPDATE averia_events
                    SET model = COALESCE(?, model), fecha = COALESCE(?, fecha),
                        inicio = COALESCE(?, inicio), fin = COALESCE(?, fin),
                        duracion_min = COALESCE(?, duracion_min), sistema = COALESCE(?, sistema),
                        tipo = COALESCE(?, tipo), descripcion = COALESCE(?, descripcion),
                        ot = COALESCE(?, ot), estado = COALESCE(?, estado),
                        severidad = ?, extra = COALESCE(?, extra)
                    WHERE id = ?
                    """,
                    (
                        event["model"], event["fecha"], event["inicio"], event["fin"],
                        event["duracion_min"], event["sistema"], event["tipo"],
                        event["descripcion"], event["ot"], event["estado"],
                        event["severidad"], event["extra"], event["id"],
                    ),
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO averia_events
                    (id, equipment_id, model, fecha, inicio, fin, duracion_min, sistema, tipo,
                     descripcion, ot, estado, severidad, origen, source_file, imported_at, extra)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        event["id"], event["equipment_id"], event["model"], event["fecha"],
                        event["inicio"], event["fin"], event["duracion_min"], event["sistema"],
                        event["tipo"], event["descripcion"], event["ot"], event["estado"],
                        event["severidad"], origen, filename, now, event["extra"],
                    ),
                )
                imported += 1
        file_hash = hashlib.sha1(data).hexdigest()
        conn.execute(
            "INSERT OR REPLACE INTO import_log (id, kind, source, filename, events, imported_at) VALUES (?, ?, ?, ?, ?, ?)",
            (file_hash, "file", origen, filename, imported, now),
        )
    return {
        "status": "ok",
        "filename": filename,
        "origen": origen,
        "events_parsed": len(events),
        "events_imported": imported,
        "events_duplicated": len(events) - imported,
        "notes": notes,
        "generated_at": now,
    }


def _decode_mime(value: str | None) -> str:
    if not value:
        return ""
    parts = decode_header(value)
    decoded = ""
    for text, charset in parts:
        if isinstance(text, bytes):
            decoded += text.decode(charset or "utf-8", errors="replace")
        else:
            decoded += text
    return decoded


def _mail_config() -> dict[str, Any]:
    return {
        "host": os.getenv("NORTHMINE_MAIL_HOST", "outlook.office365.com"),
        "user": os.getenv("NORTHMINE_MAIL_USER", ""),
        "password": os.getenv("NORTHMINE_MAIL_PASSWORD", ""),
        "folder": os.getenv("NORTHMINE_MAIL_FOLDER", "INBOX"),
        "sender": os.getenv("NORTHMINE_MAIL_SENDER", ""),
        "subject": os.getenv("NORTHMINE_MAIL_SUBJECT", ""),
        "days": int(os.getenv("NORTHMINE_MAIL_DAYS", "14")),
    }


def sync_mailbox() -> dict[str, Any]:
    config = _mail_config()
    if not config["user"] or not config["password"]:
        return {
            "status": "sin_configurar",
            "detail": "Define NORTHMINE_MAIL_USER y NORTHMINE_MAIL_PASSWORD para conectar la casilla.",
            "imported_files": [],
        }
    since = (datetime.now() - timedelta(days=config["days"])).strftime("%d-%b-%Y")
    criteria = [f'(SINCE "{since}")']
    if config["sender"]:
        criteria.append(f'(FROM "{config["sender"]}")')
    if config["subject"]:
        criteria.append(f'(SUBJECT "{config["subject"]}")')
    imported_files: list[dict[str, Any]] = []
    errors: list[str] = []
    try:
        mail = imaplib.IMAP4_SSL(config["host"])
        mail.login(config["user"], config["password"])
        mail.select(config["folder"], readonly=True)
        status, data = mail.search(None, *criteria)
        message_ids = data[0].split() if status == "OK" and data and data[0] else []
        with _connect() as conn:
            seen = {row["id"] for row in conn.execute("SELECT id FROM import_log WHERE kind = 'mail'")}
        for message_id in message_ids[-100:]:
            status, msg_data = mail.fetch(message_id, "(RFC822)")
            if status != "OK" or not msg_data or not msg_data[0]:
                continue
            message = email.message_from_bytes(msg_data[0][1])
            mail_uid = _decode_mime(message.get("Message-ID")) or message_id.decode()
            log_id = hashlib.sha1(mail_uid.encode()).hexdigest()
            if log_id in seen:
                continue
            subject = _decode_mime(message.get("Subject"))
            total_events = 0
            for part in message.walk():
                attachment_name = _decode_mime(part.get_filename())
                if not attachment_name or not attachment_name.lower().endswith(_ALLOWED_WORKBOOK_EXTENSIONS):
                    continue
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                try:
                    result = import_workbook_bytes(payload, attachment_name, origen="correo")
                    total_events += result["events_imported"]
                    imported_files.append({"filename": attachment_name, "subject": subject, **result})
                except ValueError as exc:
                    errors.append(f"{attachment_name}: {exc}")
            with _connect() as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO import_log (id, kind, source, filename, events, imported_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (log_id, "mail", subject, None, total_events, datetime.now().isoformat(timespec="seconds")),
                )
        mail.logout()
    except imaplib.IMAP4.error as exc:
        return {
            "status": "error",
            "detail": (
                f"IMAP fallo contra {config['host']}: {exc}. Si la casilla es Office365 corporativa, "
                "es probable que IMAP basico este deshabilitado por IT (requiere OAuth o app password)."
            ),
            "imported_files": imported_files,
            "errors": errors,
        }
    except OSError as exc:
        return {"status": "error", "detail": f"No se pudo conectar a {config['host']}: {exc}", "imported_files": []}
    return {"status": "ok", "imported_files": imported_files, "errors": errors}


def sync_watch_folder() -> dict[str, Any]:
    watch_dir = os.getenv("NORTHMINE_AVERIAS_WATCH_DIR", "")
    if not watch_dir:
        return {"status": "sin_configurar", "imported_files": []}
    folder = Path(watch_dir)
    if not folder.is_dir():
        return {"status": "error", "detail": f"Carpeta vigilada no existe: {watch_dir}", "imported_files": []}
    with _connect() as conn:
        seen = {row["id"] for row in conn.execute("SELECT id FROM import_log WHERE kind = 'file'")}
    imported_files = []
    errors = []
    for path in sorted(folder.glob("*.xls*")):
        if not path.is_file() or path.stat().st_size > MAX_WORKBOOK_BYTES:
            errors.append(f"{path.name}: supera el limite de {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB")
            continue
        data = path.read_bytes()
        if hashlib.sha1(data).hexdigest() in seen:
            continue
        try:
            imported_files.append(import_workbook_bytes(data, path.name, origen="carpeta"))
        except ValueError as exc:
            errors.append(f"{path.name}: {exc}")
    return {"status": "ok", "imported_files": imported_files, "errors": errors}


def _sync_all_unlocked() -> dict[str, Any]:
    folder_result = sync_watch_folder()
    mail_result = sync_mailbox()
    return {
        "folder": folder_result,
        "mail": mail_result,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def sync_all() -> dict[str, Any]:
    """Synchronize once, refusing overlapping manual/automatic executions."""
    with sync_lock("averias", ttl_seconds=900):
        return _sync_all_unlocked()


# --- Sincronizacion automatica en segundo plano ----------------------------
# Hilo daemon que ejecuta sync_all() cada NORTHMINE_AVERIAS_SYNC_MINUTES
# (default 10 min): importa lo nuevo de la carpeta vigilada y de la casilla
# IMAP si esta configurada. Se arranca desde el startup de FastAPI.

_auto_sync_state: dict[str, Any] = {"enabled": False, "interval_min": 0, "last_run": None, "last_result": None}
_auto_sync_started = False


def _auto_sync_loop(interval_seconds: float) -> None:
    import time

    while True:
        try:
            result = sync_all()
            imported = sum(
                len(result.get(source, {}).get("imported_files", []) or [])
                for source in ("folder", "mail")
            )
            _auto_sync_state["last_run"] = datetime.now().isoformat(timespec="seconds")
            _auto_sync_state["last_result"] = f"{imported} archivos nuevos"
        except Exception as exc:  # noqa: BLE001 - el loop no debe morir nunca
            _auto_sync_state["last_run"] = datetime.now().isoformat(timespec="seconds")
            _auto_sync_state["last_result"] = f"error: {exc}"
        time.sleep(interval_seconds)


def start_auto_sync() -> None:
    global _auto_sync_started
    if _auto_sync_started:
        return
    interval_min = float(os.getenv("NORTHMINE_AVERIAS_SYNC_MINUTES", "10"))
    if interval_min <= 0:
        return
    import threading

    thread = threading.Thread(target=_auto_sync_loop, args=(interval_min * 60,), daemon=True, name="averias-auto-sync")
    thread.start()
    _auto_sync_started = True
    _auto_sync_state["enabled"] = True
    _auto_sync_state["interval_min"] = interval_min


def get_mail_status() -> dict[str, Any]:
    config = _mail_config()
    with _connect() as conn:
        last_imports = [
            dict(row)
            for row in conn.execute(
                "SELECT kind, source, filename, events, imported_at FROM import_log ORDER BY imported_at DESC LIMIT 10"
            )
        ]
        total_events = conn.execute("SELECT COUNT(*) AS n FROM averia_events").fetchone()["n"]
    return {
        "mail_configured": bool(config["user"] and config["password"]),
        "mail_host": config["host"],
        "mail_user": config["user"] or None,
        "mail_folder": config["folder"],
        "mail_sender_filter": config["sender"] or None,
        "watch_dir": os.getenv("NORTHMINE_AVERIAS_WATCH_DIR") or None,
        "total_events": total_events,
        "last_imports": last_imports,
        "auto_sync": dict(_auto_sync_state),
    }


def get_latest_fleet_estado() -> dict[str, Any]:
    """Estado OP/FS por equipo segun el ultimo dia del reporte ESTATUS.

    Cubre toda la flota (perforadoras, bulldozers, aljibes, petroleros...),
    no solo los CAEX que ve WENCO. Se usa para autocompletar el checklist
    de cierre de turno del modulo Flota.
    """
    with _connect() as conn:
        last = conn.execute("SELECT MAX(fecha) AS f FROM averia_events").fetchone()["f"]
        if not last:
            return {"report_date": None, "items": []}
        rows = [
            dict(row)
            for row in conn.execute("SELECT * FROM averia_events WHERE fecha = ?", (last,))
        ]
    por_equipo: dict[str, dict[str, Any]] = {}
    for row in rows:
        estado = str(row.get("estado") or "OP").upper().replace("/", "")
        entry = por_equipo.setdefault(
            row["equipment_id"],
            {"equipment_id": row["equipment_id"], "model": row.get("model"), "estado": "OP",
             "tipo": None, "sistema": None, "descripcion": None},
        )
        # Si algun evento del dia lo deja F/S, el equipo esta F/S.
        if estado == "FS" and entry["estado"] != "FS":
            entry["estado"] = "FS"
            entry["tipo"] = row.get("tipo")
            entry["sistema"] = row.get("sistema")
            entry["descripcion"] = row.get("descripcion")
        elif entry["estado"] == "OP" and entry["tipo"] is None:
            entry["tipo"] = row.get("tipo")
            entry["sistema"] = row.get("sistema")
    return {"report_date": last, "items": sorted(por_equipo.values(), key=lambda item: item["equipment_id"])}


def _float_env(name: str, default: float) -> float:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def _downtime_cost_usd_per_min() -> float:
    # Mismo supuesto que usa Hidden Loss Detector (Decision Cockpit) para
    # dolarizar minutos de indisponibilidad: una sola tasa de referencia en
    # todo NORTHMINE, ajustable via NORTHMINE_DELAY_COST_USD_PER_MIN.
    return _float_env("NORTHMINE_DELAY_COST_USD_PER_MIN", 85.0)


def get_fleet_breakdown(days: int = 31) -> dict[str, Any]:
    cutoff = (datetime.now() - timedelta(days=max(days, 1))).date().isoformat()
    with _connect() as conn:
        rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT * FROM averia_events
                WHERE fecha IS NULL OR fecha >= ?
                ORDER BY COALESCE(inicio, fecha) DESC
                """,
                (cutoff,),
            )
        ]
    equipment: dict[str, dict[str, Any]] = {}
    total_min = 0.0
    for row in rows:
        entry = equipment.setdefault(
            row["equipment_id"],
            {
                "equipment_id": row["equipment_id"],
                "model": row["model"],
                "events": 0,
                "total_min": 0.0,
                "criticas": 0,
                "sistemas": {},
                "tipos": {},
                "last_event": None,
            },
        )
        entry["events"] += 1
        if row["duracion_min"]:
            entry["total_min"] += row["duracion_min"]
            total_min += row["duracion_min"]
        if row["severidad"] == "CRITICA":
            entry["criticas"] += 1
        if not entry["model"] and row["model"]:
            entry["model"] = row["model"]
        sistema = row["sistema"] or "Sin sistema"
        bucket = entry["sistemas"].setdefault(sistema, {"name": sistema, "events": 0, "min": 0.0})
        bucket["events"] += 1
        if row["duracion_min"]:
            bucket["min"] += row["duracion_min"]
        tipo = row["tipo"] or "Sin tipo"
        tipo_bucket = entry["tipos"].setdefault(tipo, {"name": tipo, "events": 0, "min": 0.0})
        tipo_bucket["events"] += 1
        if row["duracion_min"]:
            tipo_bucket["min"] += row["duracion_min"]
        stamp = row["inicio"] or row["fecha"]
        if stamp and (entry["last_event"] is None or stamp > entry["last_event"]):
            entry["last_event"] = stamp
    # KPIs de mantencion sobre los dias con reporte observados en la ventana.
    dias_observados = len({row["fecha"] for row in rows if row["fecha"]}) or 1
    minutos_periodo = dias_observados * 1440

    def _es_correctiva(tipo: str | None) -> bool:
        return "aver" in (tipo or "").lower()

    def _es_programada(tipo: str | None) -> bool:
        lower = (tipo or "").lower()
        return "planificado" in lower or "programada" in lower or "mantenimiento" in lower

    correctiva_min = sum(r["duracion_min"] or 0 for r in rows if _es_correctiva(r["tipo"]))
    programada_min = sum(r["duracion_min"] or 0 for r in rows if _es_programada(r["tipo"]))
    num_averias = sum(1 for r in rows if _es_correctiva(r["tipo"]))
    averias_con_duracion = [r["duracion_min"] for r in rows if _es_correctiva(r["tipo"]) and r["duracion_min"]]

    downtime_rate = _downtime_cost_usd_per_min()
    items = []
    for entry in equipment.values():
        entry["total_min"] = round(entry["total_min"], 1)
        entry["costo_usd"] = round(entry["total_min"] * downtime_rate)
        entry["disponibilidad_pct"] = round(max(0.0, (1 - entry["total_min"] / minutos_periodo)) * 100, 1)
        entry["sistemas"] = sorted(
            (
                {**bucket, "min": round(bucket["min"], 1)}
                for bucket in entry["sistemas"].values()
            ),
            key=lambda bucket: bucket["min"],
            reverse=True,
        )
        entry["tipos"] = sorted(
            (
                {**bucket, "min": round(bucket["min"], 1)}
                for bucket in entry["tipos"].values()
            ),
            key=lambda bucket: bucket["min"],
            reverse=True,
        )
        entry["tipo_principal"] = entry["tipos"][0]["name"] if entry["tipos"] else None
        items.append(entry)
    # Agrupado por modelo de equipo; dentro del modelo, el mas castigado primero.
    items.sort(key=lambda entry: ((entry["model"] or "ZZZ").upper(), -entry["total_min"]))

    # KPIs flota (sobre equipos con eventos registrados en el reporte):
    #  - DF: 1 - horas detenidas / horas calendario del periodo observado.
    #  - MTTR: duracion media de las averias correctivas.
    #  - MTBF: horas operativas / numero de averias.
    #  - % programada / correctiva: composicion de la detencion total.
    equipos_n = max(len(items), 1)
    horas_calendario = equipos_n * dias_observados * 24
    horas_detenidas = total_min / 60
    mttr_horas = round((sum(averias_con_duracion) / len(averias_con_duracion)) / 60, 1) if averias_con_duracion else None
    mtbf_horas = round((horas_calendario - horas_detenidas) / num_averias, 1) if num_averias else None
    peor = min(items, key=lambda entry: entry["disponibilidad_pct"]) if items else None
    kpis = {
        "dias_observados": dias_observados,
        "equipos": len(items),
        "disponibilidad_fisica_pct": round(max(0.0, (1 - horas_detenidas / max(horas_calendario, 1))) * 100, 1),
        "mttr_horas": mttr_horas,
        "mtbf_horas": mtbf_horas,
        "num_averias": num_averias,
        "horas_correctiva": round(correctiva_min / 60, 1),
        "horas_programada": round(programada_min / 60, 1),
        "pct_programada": round(programada_min / max(total_min, 1) * 100, 1),
        "pct_correctiva": round(correctiva_min / max(total_min, 1) * 100, 1),
        "peor_equipo": (
            {"equipment_id": peor["equipment_id"], "disponibilidad_pct": peor["disponibilidad_pct"]}
            if peor else None
        ),
    }

    costos = {
        "tasa_usd_por_min": downtime_rate,
        "total_usd": round(total_min * downtime_rate),
        "correctiva_usd": round(correctiva_min * downtime_rate),
        "programada_usd": round(programada_min * downtime_rate),
        "fuente_tasa": "NORTHMINE_DELAY_COST_USD_PER_MIN (mismo supuesto de Hidden Loss Detector)",
        "nota": "Estimado sobre minutos detenidos totales del periodo; no distingue horario productivo vs no productivo.",
    }

    return {
        "source": "xls_import",
        "days": days,
        "count_events": len(rows),
        "count_equipment": len(items),
        "total_min": round(total_min, 1),
        "kpis": kpis,
        "costos": costos,
        "equipment": items,
        # Tope amplio: con 6+ meses de historial los graficos y el analisis de
        # patrones necesitan ver todos los eventos de la ventana.
        "events": rows[:4000],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
